#!/usr/bin/env python3
"""Read an Excel worksheet and publish one JSON message per data row to OCI Streams."""

from __future__ import annotations

import argparse
import base64
import datetime as dt
import json
import os
import sys
from collections import Counter
from pathlib import Path
from typing import Any, Iterator, Sequence

from openpyxl import load_workbook

DEFAULT_REQUEST_BYTES = 900 * 1024
ONE_MIB = 1024 * 1024


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Publish Excel rows to OCI Streams as UTF-8 JSON messages."
    )
    parser.add_argument("excel", type=Path, help="Path to the source .xlsx file")
    parser.add_argument("--sheet", help="Worksheet name; defaults to the active sheet")
    parser.add_argument("--stream-ocid", default=os.getenv("OCI_STREAM_OCID"))
    parser.add_argument("--endpoint", default=os.getenv("OCI_MESSAGE_ENDPOINT"))
    parser.add_argument(
        "--config-file",
        default=os.getenv("OCI_CONFIG_FILE", "~/.oci/config"),
        help="OCI SDK config file (default: ~/.oci/config)",
    )
    parser.add_argument(
        "--profile",
        default=os.getenv("OCI_CONFIG_PROFILE", "DEFAULT"),
        help="OCI SDK profile name (default: DEFAULT)",
    )
    parser.add_argument(
        "--auth",
        choices=("config", "instance-principal"),
        default=os.getenv("OCI_AUTH", "config"),
    )
    parser.add_argument(
        "--key-columns",
        default="port_guid",
        help="Comma-separated columns used as the message key (default: port_guid)",
    )
    parser.add_argument(
        "--batch-max-bytes",
        type=int,
        default=DEFAULT_REQUEST_BYTES,
        help="Maximum decoded key+value bytes per request (default: 921600)",
    )
    parser.add_argument("--batch-max-messages", type=int, default=100)
    parser.add_argument(
        "--start-row",
        type=int,
        default=2,
        help="First Excel data row to publish, including the header offset (default: 2)",
    )
    parser.add_argument("--max-rows", type=int, help="Publish at most this many rows")
    parser.add_argument(
        "--enable-retries",
        action="store_true",
        help="Enable OCI SDK default retries; a timeout can cause duplicate messages",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Read, serialize, and batch the workbook without connecting to OCI",
    )
    return parser.parse_args()


def json_default(value: Any) -> Any:
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat()
    return str(value)


def compact_json(value: Any) -> bytes:
    return json.dumps(
        value,
        ensure_ascii=False,
        separators=(",", ":"),
        default=json_default,
    ).encode("utf-8")


def validate_headers(raw_headers: Sequence[Any]) -> list[str]:
    headers = [str(value).strip() if value is not None else "" for value in raw_headers]
    blank_positions = [index + 1 for index, value in enumerate(headers) if not value]
    duplicates = sorted(name for name, count in Counter(headers).items() if count > 1)
    if blank_positions:
        raise ValueError(f"Blank header cells at column positions: {blank_positions}")
    if duplicates:
        raise ValueError(f"Duplicate column names cannot form a JSON object: {duplicates}")
    return headers


def iter_messages(
    excel_path: Path,
    sheet_name: str | None,
    key_columns: Sequence[str],
    start_row: int,
    max_rows: int | None,
) -> tuple[str, int, Iterator[tuple[int, bytes, bytes]]]:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active
    rows = worksheet.iter_rows(values_only=True)
    try:
        headers = validate_headers(next(rows))
    except StopIteration as exc:
        workbook.close()
        raise ValueError("The worksheet is empty") from exc

    missing = [column for column in key_columns if column not in headers]
    if missing:
        workbook.close()
        raise ValueError(f"Key columns not found in worksheet: {missing}")

    def generate() -> Iterator[tuple[int, bytes, bytes]]:
        emitted = 0
        try:
            for excel_row, values in enumerate(rows, start=2):
                if excel_row < start_row:
                    continue
                if max_rows is not None and emitted >= max_rows:
                    break
                if all(value is None for value in values):
                    continue

                data = dict(zip(headers, values))
                key_parts = [
                    "" if data[column] is None else str(data[column])
                    for column in key_columns
                ]
                key = "|".join(key_parts) or f"excel-row-{excel_row}"
                payload = {
                    "schema_version": 1,
                    "source": {
                        "file": excel_path.name,
                        "sheet": worksheet.title,
                        "excel_row": excel_row,
                    },
                    "data": data,
                }
                yield excel_row, key.encode("utf-8"), compact_json(payload)
                emitted += 1
        finally:
            workbook.close()

    return worksheet.title, len(headers), generate()


def iter_batches(
    messages: Iterator[tuple[int, bytes, bytes]],
    max_bytes: int,
    max_messages: int,
) -> Iterator[list[tuple[int, bytes, bytes]]]:
    if not 1 <= max_bytes <= ONE_MIB:
        raise ValueError("--batch-max-bytes must be between 1 and 1048576")
    if max_messages < 1:
        raise ValueError("--batch-max-messages must be at least 1")

    batch: list[tuple[int, bytes, bytes]] = []
    batch_bytes = 0
    for message in messages:
        row_number, key, value = message
        message_bytes = len(key) + len(value)
        if message_bytes > ONE_MIB:
            raise ValueError(
                f"Excel row {row_number} is {message_bytes} decoded bytes, "
                "which exceeds OCI Streams' 1 MiB message limit"
            )
        if batch and (
            batch_bytes + message_bytes > max_bytes or len(batch) >= max_messages
        ):
            yield batch
            batch = []
            batch_bytes = 0
        batch.append(message)
        batch_bytes += message_bytes
    if batch:
        yield batch


def create_stream_client(args: argparse.Namespace):
    try:
        import oci
    except ImportError as exc:
        raise RuntimeError("Missing dependency: install packages from requirements.txt") from exc

    if args.auth == "instance-principal":
        signer = oci.auth.signers.InstancePrincipalsSecurityTokenSigner()
        client = oci.streaming.StreamClient(
            {}, service_endpoint=args.endpoint, signer=signer
        )
    else:
        config = oci.config.from_file(
            file_location=str(Path(args.config_file).expanduser()),
            profile_name=args.profile,
        )
        client = oci.streaming.StreamClient(config, service_endpoint=args.endpoint)
    return oci, client


def send_batch(oci: Any, client: Any, stream_ocid: str, batch, enable_retries: bool):
    entries = [
        oci.streaming.models.PutMessagesDetailsEntry(
            key=base64.b64encode(key).decode("ascii"),
            value=base64.b64encode(value).decode("ascii"),
        )
        for _, key, value in batch
    ]
    details = oci.streaming.models.PutMessagesDetails(messages=entries)
    kwargs = {}
    if enable_retries:
        kwargs["retry_strategy"] = oci.retry.DEFAULT_RETRY_STRATEGY
    response = client.put_messages(stream_ocid, details, **kwargs)

    failures = []
    for source, result in zip(batch, response.data.entries):
        row_number = source[0]
        if result.error:
            failures.append(
                f"row={row_number}, error={result.error}, message={result.error_message}"
            )
        else:
            print(
                f"published row={row_number} partition={result.partition} "
                f"offset={result.offset}",
                file=sys.stderr,
            )
    if failures:
        raise RuntimeError("OCI rejected messages: " + "; ".join(failures))


def main() -> int:
    args = parse_args()
    if args.start_row < 2:
        raise ValueError("--start-row must be 2 or greater")
    if args.max_rows is not None and args.max_rows < 1:
        raise ValueError("--max-rows must be at least 1")
    if not args.excel.is_file():
        raise FileNotFoundError(args.excel)
    if not args.dry_run and (not args.stream_ocid or not args.endpoint):
        raise ValueError(
            "Set --stream-ocid/OCI_STREAM_OCID and "
            "--endpoint/OCI_MESSAGE_ENDPOINT"
        )

    key_columns = [value.strip() for value in args.key_columns.split(",") if value.strip()]
    if not key_columns:
        raise ValueError("--key-columns must contain at least one column")

    sheet, column_count, messages = iter_messages(
        args.excel.resolve(),
        args.sheet,
        key_columns,
        args.start_row,
        args.max_rows,
    )
    batches = iter_batches(
        messages, args.batch_max_bytes, args.batch_max_messages
    )

    oci = client = None
    if not args.dry_run:
        oci, client = create_stream_client(args)

    total_rows = 0
    total_bytes = 0
    total_batches = 0
    for batch in batches:
        decoded_bytes = sum(len(key) + len(value) for _, key, value in batch)
        total_batches += 1
        total_rows += len(batch)
        total_bytes += decoded_bytes
        print(
            f"batch={total_batches} rows={len(batch)} decoded_bytes={decoded_bytes}",
            file=sys.stderr,
        )
        if not args.dry_run:
            send_batch(
                oci, client, args.stream_ocid, batch, args.enable_retries
            )

    action = "validated" if args.dry_run else "published"
    print(
        f"{action}: file={args.excel.name} sheet={sheet!r} columns={column_count} "
        f"rows={total_rows} batches={total_batches} decoded_bytes={total_bytes}",
        file=sys.stderr,
    )
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except KeyboardInterrupt:
        print("Interrupted", file=sys.stderr)
        raise SystemExit(130)
    except Exception as error:
        print(f"ERROR: {error}", file=sys.stderr)
        raise SystemExit(1)
